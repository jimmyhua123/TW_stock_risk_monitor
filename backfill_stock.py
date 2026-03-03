#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
個股多日歷史資料回補程式
用途：追溯並補齊特定股票過去多日的籌碼資料，並寫入獨立報表或合併至每日清單

優化邏輯：先一次抓取 (N + 4) 個開市日的原始資料，
再計算每天的當日數值與5日滾動累計，避免重複請求。
"""

import os
import argparse
import time
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

from risk_monitor import get_trading_date
from risk_monitor_history import get_previous_trading_days
from stock_monitor import StockDataFetcher


# ─────────────────────────────────────────────
#  Step 1: 取得所有需要的候選交易日
# ─────────────────────────────────────────────

def get_valid_dates_batch(base_date_str: str, num_days: int, window: int = 5) -> list:
    """
    取得往前推算、已驗證有開市資料的交易日清單。

    會抓取 (num_days + window - 1) 個有效交易日：
      - 最後 num_days 個是「目標輸出日」
      - 前面 (window - 1) 個是計算5日滾動所需的「前綴日」

    回傳：已驗證的交易日清單，舊→新排序。
    """
    need = num_days + (window - 1)
    # get_previous_trading_days: 包含 base_date 本身，舊→新排序
    # 我們要的是 <= base_date，加大緩衝以應對連假
    candidates = [
        d for d in get_previous_trading_days(base_date_str, need, buffer_days=20)
        if d <= base_date_str
    ]
    candidates.reverse()  # 轉為新→舊，從最近的開始驗證

    print(f"[INFO] 開始驗證開市日，需要 {need} 個有效交易日...")
    valid = []
    for date in candidates:
        if len(valid) >= need:
            break
        fetcher = StockDataFetcher(date)
        prices = fetcher.fetch_stock_prices()
        if prices:
            valid.append(date)
            print(f"  [OK] {date} (已驗證 {len(valid)}/{need})")
        else:
            print(f"  [--] {date} 無資料（可能是假日），跳過")
            time.sleep(0.5)

    valid.reverse()  # 還原為舊→新
    return valid


# ─────────────────────────────────────────────
#  Step 2: 批量一次抓取所有日期的原始資料
# ─────────────────────────────────────────────

def fetch_raw_data_for_all_dates(all_dates: list, codes: list) -> dict:
    """
    對 all_dates 中每一天都抓取：
      - 股價 (prices)
      - 三大法人 (institutional)
      - 融資融券 (margin)
    只保留 codes 中股票的資料以節省記憶體。

    回傳：{ 'YYYYMMDD': { 'prices': {...}, 'institutional': {...}, 'margin': {...} } }
    """
    raw = {}
    total = len(all_dates)
    for i, date in enumerate(all_dates):
        print(f"[{i+1}/{total}] 抓取 {date} 資料...")
        fetcher = StockDataFetcher(date)

        prices_all = fetcher.fetch_stock_prices()
        inst_all   = fetcher.fetch_institutional_trading()
        margin_all = fetcher.fetch_margin_trading()

        # 只保留目標股票
        raw[date] = {
            'prices':        {c: prices_all.get(c, {}) for c in codes},
            'institutional': {c: inst_all.get(c, {})   for c in codes},
            'margin':        {c: margin_all.get(c, {})  for c in codes},
        }
        if i < total - 1:
            time.sleep(1.5)

    return raw


# ─────────────────────────────────────────────
#  Step 3: 從快取計算每天的完整個股資料
# ─────────────────────────────────────────────

def compute_stock_data_for_date(
    target_date: str,
    all_valid_dates: list,
    raw: dict,
    code: str,
    name: str,
    window: int = 5
) -> dict:
    """
    給定 target_date，利用快取的 raw 資料計算：
      - 當日數值（價格、法人、融資）
      - 滾動 window 日累計（含當日）
    """
    idx = all_valid_dates.index(target_date)
    # 取出 window 天的日期：target_date 往前算 (window-1) 天
    window_dates = all_valid_dates[max(0, idx - window + 1): idx + 1]

    price_data = raw[target_date]['prices'].get(code, {})
    inst_data  = raw[target_date]['institutional'].get(code, {})
    margin_data = raw[target_date]['margin'].get(code, {})

    # 當日三大法人 (股→張)
    foreign_daily = round(inst_data.get('foreign_net', 0) / 1000, 0) if inst_data.get('foreign_net') is not None else None
    trust_daily   = round(inst_data.get('trust_net', 0) / 1000, 0)   if inst_data.get('trust_net') is not None else None
    dealer_daily  = round(inst_data.get('dealer_net', 0) / 1000, 0)  if inst_data.get('dealer_net') is not None else None

    # window 日累計
    foreign_wd = 0
    trust_wd   = 0
    dealer_wd  = 0
    margin_wd  = 0
    for d in window_dates:
        di = raw[d]['institutional'].get(code, {})
        dm = raw[d]['margin'].get(code, {})
        foreign_wd += di.get('foreign_net', 0) / 1000 if di.get('foreign_net') else 0
        trust_wd   += di.get('trust_net', 0) / 1000   if di.get('trust_net')   else 0
        dealer_wd  += di.get('dealer_net', 0) / 1000  if di.get('dealer_net')  else 0
        margin_wd  += dm.get('margin_change', 0)       if dm.get('margin_change') else 0

    # MA20 用快取的收盤價（取最近20個）
    close = price_data.get('close')
    ma20 = None
    if idx >= 0:
        recent_closes = [
            raw[d]['prices'].get(code, {}).get('close', 0)
            for d in all_valid_dates[max(0, idx - 19): idx + 1]
            if raw[d]['prices'].get(code, {}).get('close')
        ]
        if len(recent_closes) >= 5:
            ma20 = sum(recent_closes) / len(recent_closes)

    dist_ma20 = None
    if close and ma20:
        dist_ma20 = round((close - ma20) / ma20 * 100, 2)

    return {
        'code': code,
        'name': name,
        'close': close,
        'pct_change': price_data.get('pct_change'),
        'volume': price_data.get('volume'),
        'foreign_daily':      foreign_daily,
        'foreign_5d_sum':     round(foreign_wd, 0),
        'trust_daily':        trust_daily,
        'trust_5d_sum':       round(trust_wd, 0),
        'dealer_daily':       dealer_daily,
        'margin_daily_change': margin_data.get('margin_change'),
        'margin_5d_sum':      round(margin_wd, 0),
        'lending_daily_change': margin_data.get('short_change'),
        'dist_ma20':          dist_ma20,
    }


# ─────────────────────────────────────────────
#  Excel 工具
# ─────────────────────────────────────────────

def _chips_score(data: dict) -> str:
    score = 0
    if data.get('foreign_5d_sum'):
        score += 2 if data['foreign_5d_sum'] > 0 else -2
    if data.get('trust_5d_sum'):
        score += 1 if data['trust_5d_sum'] > 0 else -1
    if data.get('margin_5d_sum'):
        score += 1 if data['margin_5d_sum'] < 0 else -1
    if score >= 3:  return "主力積極買進"
    if score >= 1:  return "偏多"
    if score <= -3: return "主力積極賣出"
    if score <= -1: return "偏空"
    return "中性"

HEADERS = [
    '股票代號', '股票名稱', '收盤價', '漲跌幅(%)', '成交量(張)',
    '外資當日(張)', '外資5日累計', '投信當日(張)', '投信5日累計', '自營商當日(張)',
    '融資增減(張)', '融資5日累計', '借券增減(張)', 'MA20乖離(%)', '籌碼評價'
]
COL_WIDTHS = {
    '股票代號': 10, '股票名稱': 12, '收盤價': 10, '漲跌幅(%)': 10, '成交量(張)': 12,
    '外資當日(張)': 14, '外資5日累計': 14, '投信當日(張)': 14, '投信5日累計': 14,
    '自營商當日(張)': 14, '融資增減(張)': 12, '融資5日累計': 12,
    '借券增減(張)': 12, 'MA20乖離(%)': 12, '籌碼評價': 12
}

def init_sheet(ws, date_str):
    """初始化空白工作表（標題 + 表頭）"""
    ws['A1'] = f"個股籌碼監控報告 - {date_str}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells(f'A1:{get_column_letter(len(HEADERS))}1')
    ws['A1'].alignment = Alignment(horizontal='center')

    fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(3, col, h)
        c.fill = fill
        c.font = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal='center')

    for col, h in enumerate(HEADERS, 1):
        ws.column_dimensions[get_column_letter(col)].width = COL_WIDTHS.get(h, 12)


def write_stock_row(ws, row: int, data: dict):
    """將一筆股票資料寫入指定列"""
    vals = [
        data.get('code'), data.get('name'), data.get('close'),
        data.get('pct_change'), data.get('volume'),
        data.get('foreign_daily'), data.get('foreign_5d_sum'),
        data.get('trust_daily'),   data.get('trust_5d_sum'),
        data.get('dealer_daily'),
        data.get('margin_daily_change'), data.get('margin_5d_sum'),
        data.get('lending_daily_change'), data.get('dist_ma20'),
        _chips_score(data),
    ]
    for col, v in enumerate(vals, 1):
        ws.cell(row, col, v)

    # 顏色：外資
    f = data.get('foreign_daily')
    if f:
        ws.cell(row, 6).font = Font(color="008000" if f > 0 else "FF0000")
    # 顏色：漲跌幅
    p = data.get('pct_change')
    if p:
        ws.cell(row, 4).font = Font(color="FF0000" if p > 0 else "008000")


def upsert_into_existing_sheet(ws, date_str, stock_data_dict: dict):
    """
    對已存在的工作表做 upsert：
      - 若 '股票代號' 欄位有此代號 → 更新那一列
      - 否則 → 追加一列
    """
    # 確認表頭是否存在
    header_row = None
    for r in range(1, min(ws.max_row + 1, 10)):
        if ws.cell(r, 1).value in ('股票代號', HEADERS[0]):
            header_row = r
            break
    if header_row is None:
        init_sheet(ws, date_str)
        header_row = 3

    # 建立 code → row 對照表
    code_row_map = {}
    for r in range(header_row + 1, ws.max_row + 1):
        v = str(ws.cell(r, 1).value or '').strip()
        if v:
            code_row_map[v] = r

    last_row = max(code_row_map.values(), default=header_row)

    for code, data in stock_data_dict.items():
        if code in code_row_map:
            row = code_row_map[code]
        else:
            last_row += 1
            row = last_row
        write_stock_row(ws, row, data)


# ─────────────────────────────────────────────
#  主程式
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description='回補個股多日籌碼資料（批量優化版）',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例:
  # 產生獨立報表（每日一個工作表）
  python backfill_stock.py --codes 2330 2317 --days 10 --date 20260302

  # 直接將資料合併進 monitor_xlsx/YYYYMMDD.xlsx
  python backfill_stock.py --codes 3138 --days 20 --date 20260302 --merge
        """
    )
    parser.add_argument('--codes', nargs='+', required=True, help='股票代號，可多個，例如: 2330 2317')
    parser.add_argument('--days',  type=int, default=5,     help='要回補的目標天數（預設5）')
    parser.add_argument('--date',  type=str, default=None,  help='基準日期 YYYYMMDD（預設今天）')
    parser.add_argument('--merge', action='store_true',     help='是否合併至 monitor_xlsx/YYYYMMDD.xlsx')
    parser.add_argument('--window', type=int, default=5,   help='滾動計算的日窗大小（預設5）')
    args = parser.parse_args()

    base_date = get_trading_date(args.date)
    window    = args.window

    # ── Step 1: 取得 (N + window - 1) 個有效交易日 ──
    all_valid_dates = get_valid_dates_batch(base_date, args.days, window)
    if len(all_valid_dates) < window:
        print(f"[ERROR] 找不到足夠的交易日（只找到 {len(all_valid_dates)} 天）")
        return

    # 最後 num_days 個是「目標日」，前面的是「前綴」（用來計算前 window-1 日）
    target_dates = all_valid_dates[-(args.days):]
    print(f"\n[INFO] 目標回補日期: {target_dates[0]} ~ {target_dates[-1]}（共 {len(target_dates)} 天）")
    print(f"[INFO] 含前綴總計抓取: {all_valid_dates[0]} ~ {all_valid_dates[-1]}（共 {len(all_valid_dates)} 天）\n")

    # ── Step 2: 一次批量抓取所有資料 ──
    print("=" * 55)
    print(" 批量抓取所有日期原始資料（每日只抓一次）")
    print("=" * 55)
    raw = fetch_raw_data_for_all_dates(all_valid_dates, args.codes)

    # ── Step 3: 計算 + 輸出 ──
    standalone_wb = Workbook()
    standalone_wb.remove(standalone_wb.active)

    for date in target_dates:
        print(f"\n[計算] {date}")
        # 組出這天的個股資料 dict
        stock_data_for_date = {}
        for code in args.codes:
            # 嘗試從 watchlist.json 取名稱（抓不到就留空）
            name = raw[date]['prices'].get(code, {}).get('name', '')
            data = compute_stock_data_for_date(date, all_valid_dates, raw, code, name, window)
            stock_data_for_date[code] = data

        if args.merge:
            file_path = os.path.join('monitor_xlsx', f"{date}.xlsx")
            if os.path.exists(file_path):
                try:
                    wb = load_workbook(file_path)
                    sheet_name = next((s for s in ('個股籌碼', '個股資料') if s in wb.sheetnames), None)
                    if sheet_name:
                        ws = wb[sheet_name]
                        upsert_into_existing_sheet(ws, date, stock_data_for_date)
                    else:
                        ws = wb.create_sheet("個股籌碼")
                        init_sheet(ws, date)
                        for i, (code, data) in enumerate(stock_data_for_date.items(), 4):
                            write_stock_row(ws, i, data)
                    wb.save(file_path)
                    print(f"  [SUCCESS] 已更新 {file_path}")
                except Exception as e:
                    print(f"  [ERROR] 無法更新 {file_path}: {e}")
            else:
                print(f"  [WARNING] 找不到 {file_path}，跳過（可不加 --merge 改用獨立輸出）")
        else:
            ws = standalone_wb.create_sheet(title=date)
            init_sheet(ws, date)
            for row_i, (code, data) in enumerate(stock_data_for_date.items(), 4):
                write_stock_row(ws, row_i, data)

    if not args.merge:
        codes_str = '_'.join(args.codes)
        output_name = f"backfill_{codes_str}_{target_dates[0]}_{target_dates[-1]}.xlsx"
        standalone_wb.save(output_name)
        print(f"\n[SUCCESS] 獨立報表已輸出: {output_name}")


if __name__ == '__main__':
    main()
