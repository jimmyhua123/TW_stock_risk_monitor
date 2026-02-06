#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣股市風險監控 - 主整合程式
整合單日數據、歷史統計與個股籌碼監控，輸出 Excel 報表
"""

import sys
import os
import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# 匯入現有模組
from risk_monitor import RiskMonitor, get_trading_date
from risk_monitor_history import HistoricalDataFetcher
from stock_monitor import StockMonitor


class IntegratedRiskReport:
    """整合風險報告生成器"""
    
    def __init__(self, date_str: str, watchlist_path: str = 'watchlist.json'):
        self.date_str = date_str
        self.watchlist_path = watchlist_path
        self.single_day_data = {}
        self.history_data = {}
        self.stock_data = {}  # 個股籌碼資料
    
    def fetch_all_data(self):
        """抓取所有數據（單日 + 歷史統計）"""
        print(f"[INFO] 開始抓取 {self.date_str} 的完整風險監控數據...\n")
        
        # 1. 抓取單日數據
        print("=" * 60)
        print(" 第一階段：單日數據抓取")
        print("=" * 60)
        monitor = RiskMonitor(self.date_str)
        monitor.fetch_all_data()
        self.single_day_data = monitor.data
        
        # 2. 抓取歷史統計
        print("\n" + "=" * 60)
        print(" 第二階段：歷史統計計算")
        print("=" * 60)
        hist_fetcher = HistoricalDataFetcher(self.date_str)
        
        # 三大法人歷史（20日）
        inst_hist = hist_fetcher.fetch_institutional_history(20)
        
        # 融資融券歷史（20日）
        margin_hist = hist_fetcher.fetch_margin_history(20)
        
        # P/C Ratio 歷史（5日）
        pc_hist = hist_fetcher.fetch_pc_ratio_history(5)
        
        # 期貨歷史（5日）
        futures_hist = hist_fetcher.fetch_futures_history(5)
        
        self.history_data = {
            'institutional': inst_hist,
            'margin': margin_hist,
            'pc_ratio': pc_hist,
            'futures': futures_hist
        }
        
        # 3. 抓取個股籌碼資料
        print("\n" + "=" * 60)
        print(" 第三階段：個股籌碼監控")
        print("=" * 60)
        try:
            stock_monitor = StockMonitor(self.date_str, self.watchlist_path)
            stock_monitor.load_watchlist()
            stock_monitor.fetch_all_data()
            self.stock_data = stock_monitor.stock_data
        except Exception as e:
            print(f"[WARNING] 個股籌碼抓取失敗: {e}")
            self.stock_data = {}
        
        print("\n[SUCCESS] 所有數據抓取完成！\n")
    
    def export_to_excel(self, filename: str = 'risk_report.xlsx'):
        """匯出到 Excel 檔案"""
        print(f"[INFO] 正在生成 Excel 報表...")
        
        wb = Workbook()
        
        # 刪除預設工作表
        wb.remove(wb.active)
        
        # 創建工作表
        ws_summary = wb.create_sheet("總覽", 0)
        ws_detail = wb.create_sheet("詳細數據", 1)
        ws_stock = wb.create_sheet("個股籌碼", 2)
        
        # 生成總覽表
        self._create_summary_sheet(ws_summary)
        
        # 生成詳細數據表
        self._create_detail_sheet(ws_detail)
        
        # 生成個股籌碼表
        self._create_stock_sheet(ws_stock)
        
        # 確保 monitor_xlsx 目錄存在
        output_dir = 'monitor_xlsx'
        os.makedirs(output_dir, exist_ok=True)
        
        # 組合完整輸出路徑
        output_path = os.path.join(output_dir, filename)
        
        # 儲存檔案
        wb.save(output_path)
        print(f"[SUCCESS] Excel 報表已儲存至: {output_path}\n")
    
    def _create_summary_sheet(self, ws):
        """創建總覽工作表"""
        # 標題
        ws['A1'] = f"台灣股市風險監控報告 - {self.date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:H1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 設定表頭
        headers = ['類別', '指標', '當日數值', '單日變動', '5日平均', '5日總和', '20日平均', '20日總和']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # 填入數據
        row = 4
        for indicator in self.single_day_data.get('indicators', []):
            name = indicator['name']
            value = indicator['value']
            change = indicator['change']
            unit = indicator['unit']
            
            # 當日數值
            value_str = f"{value}{unit}" if value is not None else "N/A"
            change_str = f"{change:+.2f}%" if change is not None else ""
            
            # 歷史統計（根據指標類型）
            stats_5d_avg, stats_5d_sum, stats_20d_avg, stats_20d_sum = "", "", "", ""
            
            if name == '外資現貨':
                stats_5d_avg = f"{self.history_data['institutional'].get('foreign_5d_avg', '-')}億"
                stats_5d_sum = f"{self.history_data['institutional'].get('foreign_5d_sum', '-')}億"
                stats_20d_avg = f"{self.history_data['institutional'].get('foreign_20d_avg', '-')}億"
                stats_20d_sum = f"{self.history_data['institutional'].get('foreign_20d_sum', '-')}億"
            elif name == '投信現貨':
                stats_5d_avg = f"{self.history_data['institutional'].get('trust_5d_avg', '-')}億"
                stats_5d_sum = f"{self.history_data['institutional'].get('trust_5d_sum', '-')}億"
            elif name == '融資融券變化':
                stats_5d_avg = f"{self.history_data['margin'].get('margin_5d_avg', '-')}億"
                stats_5d_sum = f"{self.history_data['margin'].get('margin_5d_sum', '-')}億"
                stats_20d_avg = f"{self.history_data['margin'].get('margin_20d_avg', '-')}億"
                stats_20d_sum = f"{self.history_data['margin'].get('margin_20d_sum', '-')}億"
            elif name == '選擇權 P/C Ratio':
                stats_5d_avg = f"{self.history_data['pc_ratio'].get('pc_5d_avg', '-')}%"
            elif name == '外資期貨未平倉':
                stats_5d_avg = f"{self.history_data['futures'].get('futures_5d_avg', '-')}口"
            
            # 寫入行
            ws.cell(row, 1, indicator['category'])
            ws.cell(row, 2, name)
            ws.cell(row, 3, value_str)
            ws.cell(row, 4, change_str)
            ws.cell(row, 5, stats_5d_avg)
            ws.cell(row, 6, stats_5d_sum)
            ws.cell(row, 7, stats_20d_avg)
            ws.cell(row, 8, stats_20d_sum)
            
            row += 1
        
        # 調整欄寬
        for col in range(1, 9):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_detail_sheet(self, ws):
        """創建詳細數據工作表"""
        ws['A1'] = "詳細歷史統計數據"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 三大法人區塊
        row = 3
        ws[f'A{row}'] = "三大法人買賣超（過去20日）"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        inst = self.history_data.get('institutional', {})
        data_rows = [
            ['外資 5日平均', f"{inst.get('foreign_5d_avg', '-')}億"],
            ['外資 5日總和', f"{inst.get('foreign_5d_sum', '-')}億"],
            ['外資 20日平均', f"{inst.get('foreign_20d_avg', '-')}億"],
            ['外資 20日總和', f"{inst.get('foreign_20d_sum', '-')}億"],
            ['', ''],
            ['投信 5日平均', f"{inst.get('trust_5d_avg', '-')}億"],
            ['投信 5日總和', f"{inst.get('trust_5d_sum', '-')}億"],
        ]
        
        for data_row in data_rows:
            ws.cell(row, 1, data_row[0])
            ws.cell(row, 2, data_row[1])
            row += 1
        
        # 融資融券區塊
        row += 2
        ws[f'A{row}'] = "融資融券變化（過去20日）"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        margin = self.history_data.get('margin', {})
        margin_rows = [
            ['5日平均', f"{margin.get('margin_5d_avg', '-')}億"],
            ['5日總和', f"{margin.get('margin_5d_sum', '-')}億"],
            ['20日平均', f"{margin.get('margin_20d_avg', '-')}億"],
            ['20日總和', f"{margin.get('margin_20d_sum', '-')}億"],
        ]
        
        for data_row in margin_rows:
            ws.cell(row, 1, data_row[0])
            ws.cell(row, 2, data_row[1])
            row += 1
        
        # 期貨與選擇權區塊
        row += 2
        ws[f'A{row}'] = "期貨與選擇權（過去5日）"
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
        
        pc = self.history_data.get('pc_ratio', {})
        futures = self.history_data.get('futures', {})
        
        futures_rows = [
            ['P/C Ratio 5日平均', f"{pc.get('pc_5d_avg', '-')}%"],
            ['外資期貨淨部位 5日平均', f"{futures.get('futures_5d_avg', '-')}口"],
        ]
        
        for data_row in futures_rows:
            ws.cell(row, 1, data_row[0])
            ws.cell(row, 2, data_row[1])
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _create_stock_sheet(self, ws):
        """創建個股籌碼監控工作表"""
        # 標題
        ws['A1'] = f"個股籌碼監控報告 - {self.date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        if not self.stock_data:
            ws['A3'] = "無個股籌碼資料（請確認 watchlist.json 存在）"
            return
        
        # 表頭 (包含進階籌碼指標)
        headers = [
            '股票代號', '股票名稱', '收盤價', '漲跌幅(%)', '成交量(張)',
            '外資當日(張)', '外資5日累計', '投信當日(張)', '投信5日累計', '自營商當日(張)',
            '融資增減(張)', '融資5日累計', '借券增減(張)', 'MA20乖離(%)', '籌碼評價',
            '買賣券商差', '籌碼集中度5D(%)', '借券賣出餘額', '短回補天數', 'VWAP20D', 'VWAP乖離(%)', '資料來源'
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 資料行
        row = 4
        for code, data in self.stock_data.items():
            # 計算籌碼評價
            chips_score = self._evaluate_chips(data)
            
            ws.cell(row, 1, data.get('code', ''))
            ws.cell(row, 2, data.get('name', ''))
            ws.cell(row, 3, data.get('close'))
            ws.cell(row, 4, data.get('pct_change'))
            ws.cell(row, 5, data.get('volume'))
            ws.cell(row, 6, data.get('foreign_daily'))
            ws.cell(row, 7, data.get('foreign_5d_sum'))
            ws.cell(row, 8, data.get('trust_daily'))
            ws.cell(row, 9, data.get('trust_5d_sum'))
            ws.cell(row, 10, data.get('dealer_daily'))
            ws.cell(row, 11, data.get('margin_daily_change'))
            ws.cell(row, 12, data.get('margin_5d_sum'))
            ws.cell(row, 13, data.get('lending_daily_change'))
            ws.cell(row, 14, data.get('dist_ma20'))
            ws.cell(row, 15, chips_score)
            
            # 進階籌碼指標
            ws.cell(row, 16, data.get('broker_buy_sell_diff'))
            ws.cell(row, 17, data.get('chip_concentration_5d'))
            ws.cell(row, 18, data.get('sbl_sell_balance'))
            ws.cell(row, 19, data.get('short_cover_days'))
            ws.cell(row, 20, data.get('vwap_20d_approx'))
            ws.cell(row, 21, data.get('vwap_bias'))
            ws.cell(row, 22, data.get('data_source', 'N/A'))
            
            # 條件格式 - 外資買超綠色，賣超紅色
            if data.get('foreign_daily') and data['foreign_daily'] > 0:
                ws.cell(row, 6).font = Font(color="008000")
            elif data.get('foreign_daily') and data['foreign_daily'] < 0:
                ws.cell(row, 6).font = Font(color="FF0000")
            
            # 漲跌幅顏色
            if data.get('pct_change') and data['pct_change'] > 0:
                ws.cell(row, 4).font = Font(color="FF0000")
            elif data.get('pct_change') and data['pct_change'] < 0:
                ws.cell(row, 4).font = Font(color="008000")
            
            # 籌碼集中度顏色
            if data.get('chip_concentration_5d'):
                if data['chip_concentration_5d'] > 0:
                    ws.cell(row, 17).font = Font(color="008000")
                elif data['chip_concentration_5d'] < 0:
                    ws.cell(row, 17).font = Font(color="FF0000")
            
            row += 1
        
        # 調整欄寬 (含進階指標欄位)
        col_widths = [10, 12, 10, 10, 12, 14, 14, 14, 14, 14, 12, 12, 12, 12, 14, 12, 14, 14, 12, 10, 12, 10]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def _evaluate_chips(self, data: dict) -> str:
        """評估籌碼面"""
        score = 0
        
        # 外資連續買
        if data.get('foreign_5d_sum') and data['foreign_5d_sum'] > 0:
            score += 2
        elif data.get('foreign_5d_sum') and data['foreign_5d_sum'] < 0:
            score -= 2
        
        # 投信連續買
        if data.get('trust_5d_sum') and data['trust_5d_sum'] > 0:
            score += 1
        elif data.get('trust_5d_sum') and data['trust_5d_sum'] < 0:
            score -= 1
        
        # 融資減少 (籌碼乾淨)
        if data.get('margin_5d_sum') and data['margin_5d_sum'] < 0:
            score += 1
        elif data.get('margin_5d_sum') and data['margin_5d_sum'] > 0:
            score -= 1
        
        # 評價
        if score >= 3:
            return "主力積極買進"
        elif score >= 1:
            return "偏多"
        elif score <= -3:
            return "主力積極賣出"
        elif score <= -1:
            return "偏空"
        else:
            return "中性"


def main():
    """主程式"""
    parser = argparse.ArgumentParser(
        description='台灣股市風險監控 - 整合報告生成器',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例:
  # 生成指定日期的完整報告
  python main.py --date 20260123 --output report_20260123.xlsx
  
  # 生成今天的報告（週末自動回退）
  python main.py
        """
    )
    
    parser.add_argument(
        '--date',
        type=str,
        help='指定日期 (格式: YYYYMMDD)，不指定則使用今天'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        default='risk_report.xlsx',
        help='Excel 輸出檔名（預設: risk_report.xlsx）'
    )
    

    
    args = parser.parse_args()
    
    # 取得交易日期
    trading_date = get_trading_date(args.date)
    
    if args.date and args.date != trading_date:
        print(f"[INFO] 指定日期為週末，已調整為 {trading_date}\n")
    
    # 執行整合報告生成
    try:
        report = IntegratedRiskReport(trading_date)
        report.fetch_all_data()
        report.export_to_excel(args.output)
        
        print(f"[SUCCESS] 完成！報告已儲存至 {args.output}")
        
    except KeyboardInterrupt:
        print("\n\n[WARNING] 程式被使用者中斷")
        sys.exit(1)
    except Exception as e:
        print(f"\n[ERROR] 執行失敗: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
