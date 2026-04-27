#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣個股籌碼監控程式
抓取個股三大法人買賣超、融資融券、技術面資料
"""

import requests
import pandas as pd
import json
import os
import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
from io import StringIO
import time
import re
from bs4 import BeautifulSoup

# 匯入現有模組
from risk_monitor import get_trading_date
from risk_monitor_history import get_previous_trading_days

class StockDataFetcher:
    """個股資料抓取器 (支援上市 TWSE + 上櫃 TPEx)"""
    
    TWSE_BASE_URL = "https://www.twse.com.tw"
    TPEX_BASE_URL = "https://www.tpex.org.tw"
    TAIFEX_BASE_URL = "https://www.taifex.com.tw"
    
    def __init__(self, date_str: str, price_cache: dict = None):
        """
        Args:
            date_str: 日期字串，格式 YYYYMMDD
            price_cache: 可選，預先填入的股價快取（避免重複請求）
        """
        self.date_str = date_str
        self.formatted_date = f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
        # TPEx 使用民國年格式: YYY/MM/DD
        year_ad = int(date_str[:4])
        year_roc = year_ad - 1911
        self.tpex_date = f"{year_roc}/{date_str[4:6]}/{date_str[6:]}"
        
        # 快取資料
        self._institutional_cache = None
        self._margin_cache = None
        self._price_cache = price_cache if price_cache else None
        self._warrant_master = None
    
    def fetch_institutional_trading(self) -> Dict[str, Dict[str, Any]]:
        """
        抓取三大法人個股買賣超資料 (T86)
        Returns:
            {
                '2330': {
                    'foreign_buy': 外資買進股數,
                    'foreign_sell': 外資賣出股數,
                    'foreign_net': 外資買賣超股數,
                    'trust_buy': 投信買進股數,
                    'trust_sell': 投信賣出股數,
                    'trust_net': 投信買賣超股數,
                    'dealer_buy': 自營商買進股數,
                    'dealer_sell': 自營商賣出股數,
                    'dealer_net': 自營商買賣超股數,
                },
                ...
            }
        """
        if self._institutional_cache is not None:
            return self._institutional_cache
            
        try:
            url = f"{self.TWSE_BASE_URL}/fund/T86"
            params = {
                'response': 'json',
                'date': self.date_str,
                'selectType': 'ALLBUT0999'  # 全部但排除權證
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                print(f"[WARNING] T86 API 返回錯誤: {data.get('stat')}")
                return {}
            
            result = {}
            for row in data.get('data', []):
                if len(row) >= 17:
                    code = row[0].strip()
                    name = row[1].strip()
                    
                    # 解析數值 (移除逗號)
                    def parse_int(val):
                        try:
                            return int(str(val).replace(',', ''))
                        except:
                            return 0
                    
                    # T86 欄位順序:
                    # 0: 證券代號, 1: 證券名稱
                    # 2-4: 外陸資買/賣/淨 (不含自營商)
                    # 5-7: 外資自營商買/賣/淨
                    # 8-10: 投信買/賣/淨
                    # 11-13: 自營商(自行)買/賣/淨
                    # 14-16: 自營商(避險)買/賣/淨
                    # 17: 三大法人合計買賣超
                    result[code] = {
                        'name': name,
                        'foreign_buy': parse_int(row[2]),
                        'foreign_sell': parse_int(row[3]),
                        'foreign_net': parse_int(row[4]),
                        'trust_buy': parse_int(row[8]),
                        'trust_sell': parse_int(row[9]),
                        'trust_net': parse_int(row[10]),
                        'dealer_buy': parse_int(row[11]) + parse_int(row[14]),  # 自行買賣 + 避險
                        'dealer_sell': parse_int(row[12]) + parse_int(row[15]),
                        'dealer_net': parse_int(row[13]) + parse_int(row[16]),
                    }
            
            self._institutional_cache = result
            
            # 額外抓取權證 (0999) 三大法人資料
            try:
                params['selectType'] = '0999'
                resp_w = requests.get(url, params=params, timeout=15)
                if resp_w.status_code == 200:
                    data_w = resp_w.json()
                    if data_w.get('stat') == 'OK':
                        for row in data_w.get('data', []):
                            if len(row) >= 17:
                                code = row[0].strip()
                                result[code] = {
                                    'name': row[1].strip(),
                                    'foreign_buy': parse_int(row[2]),
                                    'foreign_sell': parse_int(row[3]),
                                    'foreign_net': parse_int(row[4]),
                                    'trust_buy': parse_int(row[8]),
                                    'trust_sell': parse_int(row[9]),
                                    'trust_net': parse_int(row[10]),
                                    'dealer_buy': parse_int(row[11]) + parse_int(row[14]),
                                    'dealer_sell': parse_int(row[12]) + parse_int(row[15]),
                                    'dealer_net': parse_int(row[13]) + parse_int(row[16]),
                                }
            except Exception as e:
                print(f"[WARNING] 抓取權證三大法人資料失敗: {e}")

            # 合併 TPEx (上櫃) 資料
            tpex_data = self._fetch_tpex_institutional()
            result.update(tpex_data)
            self._institutional_cache = result
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取三大法人個股資料失敗: {e}")
            return {}
    
    def _fetch_tpex_institutional(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 三大法人買賣超資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/3insti/daily_trade/3itrade_hedge_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            tables = data.get('tables', [])
            # TPEx API: tables[0] 是 dict，資料在 tables[0]['data']
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 24:
                        code = str(row[0]).strip()
                        name = str(row[1]).strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', ''))
                            except:
                                return 0
                        
                        # TPEx 欄位順序:
                        # 0: 代號, 1: 名稱
                        # 2-4: 外資買/賣/淨 (不含自營商)
                        # 5-7: 外資自營商買/賣/淨  
                        # 8-10: 外資合計買/賣/淨
                        # 11-13: 投信買/賣/淨
                        # 14-16: 自營商(自行)買/賣/淨
                        # 17-19: 自營商(避險)買/賣/淨
                        # 20-22: 自營商合計買/賣/淨
                        # 23: 三大法人合計
                        result[code] = {
                            'name': name,
                            'foreign_buy': parse_int(row[2]),
                            'foreign_sell': parse_int(row[3]),
                            'foreign_net': parse_int(row[4]),
                            'trust_buy': parse_int(row[11]),
                            'trust_sell': parse_int(row[12]),
                            'trust_net': parse_int(row[13]),
                            'dealer_buy': parse_int(row[14]) + parse_int(row[17]),
                            'dealer_sell': parse_int(row[15]) + parse_int(row[18]),
                            'dealer_net': parse_int(row[16]) + parse_int(row[19]),
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 三大法人資料失敗: {e}")
            return {}
    
    def fetch_margin_trading(self) -> Dict[str, Dict[str, Any]]:
        """
        抓取個股融資融券餘額
        Returns:
            {
                '2330': {
                    'margin_buy': 融資買進,
                    'margin_sell': 融資賣出,
                    'margin_balance': 融資餘額,
                    'margin_change': 融資增減,
                    'short_buy': 融券買進,
                    'short_sell': 融券賣出,
                    'short_balance': 融券餘額,
                    'short_change': 融券增減,
                },
                ...
            }
        """
        if self._margin_cache is not None:
            return self._margin_cache
            
        try:
            url = f"{self.TWSE_BASE_URL}/rwd/zh/marginTrading/MI_MARGN"
            params = {
                'response': 'json',
                'date': self.date_str,
                'selectType': 'ALL'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                print(f"[WARNING] 融資融券 API 返回錯誤: {data.get('stat')}")
                return {}
            
            result = {}
            
            # 解析個股融資融券資料 (tables[1] 是個股明細)
            if len(data.get('tables', [])) > 1:
                table_data = data['tables'][1].get('data', [])
                
                for row in table_data:
                    if len(row) >= 13:
                        code = row[0].strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', '').replace('--', '0'))
                            except:
                                return 0
                        
                        result[code] = {
                            'margin_buy': parse_int(row[2]),
                            'margin_sell': parse_int(row[3]),
                            'margin_balance': parse_int(row[6]),
                            'margin_change': parse_int(row[6]) - parse_int(row[1]) if row[1] != '--' else 0,
                            'short_sell': parse_int(row[8]),
                            'short_buy': parse_int(row[9]),
                            'short_balance': parse_int(row[12]),
                            'short_change': parse_int(row[12]) - parse_int(row[7]) if row[7] != '--' else 0,
                        }
            
            self._margin_cache = result
            
            # 合併 TPEx (上櫃) 資料
            tpex_data = self._fetch_tpex_margin()
            result.update(tpex_data)
            self._margin_cache = result
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取融資融券資料失敗: {e}")
            return {}
    
    def _fetch_tpex_margin(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 融資融券資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/margin_trading/margin_balance/margin_bal_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            # TPEx API: 資料在 tables[0]['data']
            tables = data.get('tables', [])
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 7:
                        code = str(row[0]).strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', '').replace('--', '0'))
                            except:
                                return 0
                        
                        # TPEx 融資融券欄位:
                        # 0: 代號, 1: 名稱
                        # 2: 前日融資餘額, 3: 融資買進, 4: 融資賣出, 5: 現金償還, 6: 今日融資餘額
                        result[code] = {
                            'margin_buy': parse_int(row[3]) if len(row) > 3 else 0,
                            'margin_sell': parse_int(row[4]) if len(row) > 4 else 0,
                            'margin_balance': parse_int(row[6]) if len(row) > 6 else 0,
                            'margin_change': parse_int(row[6]) - parse_int(row[2]) if len(row) > 6 else 0,
                            'short_sell': 0,
                            'short_buy': 0,
                            'short_balance': 0,
                            'short_change': 0,
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 融資融券資料失敗: {e}")
            return {}

    def fetch_warrant_master(self) -> Dict[str, Dict[str, Any]]:
        """
        從證交所 OpenAPI 抓取所有權證的基本資料 (履約價、行使比例、到期日)
        """
        if self._warrant_master is not None:
            return self._warrant_master
            
        try:
            # 認購售權證基本資料 (t187ap37_L)
            url = "https://openapi.twse.com.tw/v1/opendata/t187ap37_L"
            print(f"[DEBUG] Fetching warrant master from {url}...")
            response = requests.get(url, timeout=30)
            response.raise_for_status()
            data = response.json()
            print(f"[DEBUG] Received {len(data)} warrant master items")
            
            def pick_first(item: Dict[str, Any], *keys: str, default: Any = '') -> Any:
                for key in keys:
                    if key in item and item.get(key) not in (None, ''):
                        return item.get(key)
                return default

            def parse_float(value: Any) -> float:
                try:
                    return float(str(value).replace(',', '').strip() or 0)
                except Exception:
                    return 0.0

            result = {}
            for item in data:
                code = str(pick_first(item, '權證代號')).strip()
                if not code:
                    continue
                
                # 到期日處理 (ROC year 1150803 -> 20260803)
                expiry_roc = str(
                    pick_first(item, '到期日', '履約截止日', '最後交易日')
                ).strip()
                expiry_date = ""
                if len(expiry_roc) == 7:
                    year = int(expiry_roc[:3]) + 1911
                    expiry_date = f"{year}{expiry_roc[3:]}"
                
                result[code] = {
                    'strike_price': parse_float(
                        pick_first(
                            item,
                            '最新履約價格(元)/履約指數',
                            '最新履約價(元)/指數',
                        )
                    ),
                    'exercise_ratio': parse_float(
                        pick_first(
                            item,
                            '最新標的履約配發數量(每仟單位權證)',
                            '最新標的履約量(千股/千指數)',
                            '最新標的履約量(每張權證',
                        )
                    ) / 1000.0,
                    'expiry_date': expiry_date,
                    'underlying_code': str(
                        pick_first(item, '標的代號', '標的證券/指數')
                    ).strip(),
                    'underlying_name': str(pick_first(item, '標的證券/指數')).strip(),
                }
            self._warrant_master = result
            return result
        except Exception as e:
            print(f"[WARNING] 抓取權證基本資料失敗: {e}")
            return {}

    def fetch_warrant_details(self, wid: str) -> Dict[str, Any]:
        """
        從元大權證網抓取權證詳細參數
        """
        url = f"https://www.warrantwin.com.tw/eyuanta/Warrant/Info.aspx?WID={wid}"
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36',
            'Referer': 'https://www.warrantwin.com.tw/'
        }
        
        try:
            res = requests.get(url, headers=headers, timeout=10)
            res.encoding = res.apparent_encoding # 自動偵測編碼 (通常是 utf-8 或 big5)
            soup = BeautifulSoup(res.text, 'html.parser')
            
            details = {}
            # 增加對應標籤，排除可能誤導的註解部分
            mapping = {
                '最新履約價': 'strike_price',
                '最新行使比例': 'exercise_ratio',
                '到期日期': 'expiry_date',
                '剩餘天數': 'days_to_expiry',
                '價內外程度': 'moneyness',
                '買賣價差比': 'bid_ask_spread_pct',
                '實質槓桿': 'effective_leverage',
                '買價隱波': 'implied_volatility',
                '流通在外張數/比例': 'outstanding_info'
            }
            
            # 專門尋找「基本資料」區塊或直接找關鍵字
            # 遍歷所有包含文字的標籤
            for element in soup.find_all(['span', 'li', 'td', 'div']):
                text = element.get_text(strip=True)
                for label, key in mapping.items():
                    if label in text and key not in details:
                        # 排除掉只有標籤本身的長度
                        if len(text) > len(label):
                            val = text.replace(label, '').replace(':', '').strip()
                            if val:
                                details[key] = val
            
            # 處理流通在外比例 %
            if 'outstanding_info' in details:
                match = re.search(r'([\d.]+)%', details['outstanding_info'])
                if match:
                    details['outstanding_pct'] = match.group(1)
            
            return details
        except Exception as e:
            print(f"[WARNING] 抓取權證 {wid} 詳細資料失敗: {e}")
            return {}
    
    def fetch_stock_prices(self, include_warrants: bool = True) -> Dict[str, Dict[str, Any]]:
        """
        抓取個股收盤行情
        Returns:
            {
                '2330': {
                    'close': 收盤價,
                    'change': 漲跌,
                    'pct_change': 漲跌幅 %,
                    'volume': 成交量 (張),
                },
                ...
            }
        """
        if self._price_cache is not None:
            return self._price_cache
            
        MAX_RETRIES = 3
        last_error = None
        for attempt in range(1, MAX_RETRIES + 1):
            try:
                # 使用新版 API 端點
                url = f"{self.TWSE_BASE_URL}/rwd/zh/afterTrading/MI_INDEX"
                params = {
                    'response': 'json',
                    'date': self.date_str,
                    'type': 'ALLBUT0999'
                }
                
                response = requests.get(url, params=params, timeout=30)
                response.raise_for_status()
                data = response.json()
                
                if data.get('stat') != 'OK':
                    # stat 非 OK 通常表示假日或無資料，不需重試
                    print(f"[WARNING] 股價 API 返回錯誤: {data.get('stat')}")
                    return {}
                
                result = {}
                
                # 嘗試從多個表格抓取收盤行情 (Table 8 為個股, Table 9 為權證)
                tables = data.get('tables', [])
                
                def parse_float(val):
                    try:
                        clean = str(val).replace(',', '').replace('X', '')
                        clean = clean.replace('<p style= color:green>', '').replace('<p style= color:red>', '').replace('</p>', '')
                        clean = clean.replace('+', '').strip()
                        if clean == '--' or clean == '':
                            return 0.0
                        return float(clean)
                    except:
                        return 0.0

                def parse_int(val):
                    try:
                        return int(str(val).replace(',', ''))
                    except:
                        return 0

                def process_table(table_data):
                    fields = table_data.get('fields', [])
                    field_idx = {str(name).strip(): idx for idx, name in enumerate(fields)}

                    def idx_of(*names: str) -> Optional[int]:
                        for name in names:
                            if name in field_idx:
                                return field_idx[name]
                        return None

                    code_idx = idx_of('證券代號')
                    name_idx = idx_of('證券名稱')
                    volume_idx = idx_of('成交股數')
                    close_idx = idx_of('收盤價')
                    sign_idx = idx_of('漲跌(+/-)')
                    change_idx = idx_of('漲跌價差')
                    bid_idx = idx_of('最後揭示買價')
                    ask_idx = idx_of('最後揭示賣價')
                    underlying_code_idx = idx_of('標的代號')
                    underlying_price_idx = idx_of('標的收盤價/指數')

                    if None in (code_idx, name_idx, volume_idx, close_idx, sign_idx, change_idx):
                        return

                    for row in table_data.get('data', []):
                        if len(row) <= max(code_idx, name_idx, volume_idx, close_idx, sign_idx, change_idx):
                            continue

                        code = str(row[code_idx]).strip()
                        if not code.isdigit():
                            continue

                        close = parse_float(row[close_idx])
                        change_val = parse_float(row[change_idx])
                        volume = parse_int(row[volume_idx]) // 1000  # 股數 -> 張

                        if 'green' in str(row[sign_idx]):
                            change_val = -abs(change_val)

                        pct_change = 0.0
                        if close > 0 and (close - change_val) != 0:
                            pct_change = round((change_val / (close - change_val)) * 100, 2)

                        result[code] = {
                            'name': str(row[name_idx]).strip(),
                            'market': '上市',
                            'close': close,
                            'change': change_val,
                            'pct_change': pct_change,
                            'volume': volume,
                        }

                        if bid_idx is not None and ask_idx is not None and len(row) > max(bid_idx, ask_idx):
                            result[code].update({
                                'bid': parse_float(row[bid_idx]),
                                'ask': parse_float(row[ask_idx]),
                            })

                        if underlying_code_idx is not None and len(row) > underlying_code_idx:
                            result[code]['underlying_code'] = str(row[underlying_code_idx]).strip()

                        if underlying_price_idx is not None and len(row) > underlying_price_idx:
                            result[code]['underlying_price'] = parse_float(row[underlying_price_idx])

                # 處理個股 (ALLBUT0999)
                for table in tables:
                    if '每日收盤行情' in table.get('title', ''):
                        process_table(table)
                
                # 額外抓取權證 (0999)
                if include_warrants:
                    try:
                        params['type'] = '0999'
                        print(f"[DEBUG] Fetching warrant prices for {self.date_str}...")
                        resp_w = requests.get(url, params=params, timeout=60) # 權證資料很大，給 60s
                        if resp_w.status_code == 200:
                            data_w = resp_w.json()
                            if data_w.get('stat') == 'OK':
                                w_count = 0
                                for table in data_w.get('tables', []):
                                    if '每日收盤行情' in table.get('title', ''):
                                        start_count = len(result)
                                        process_table(table)
                                        w_count += (len(result) - start_count)
                                print(f"[DEBUG] Added {w_count} warrant prices")
                    except Exception as e:
                        print(f"[WARNING] 抓取權證股價失敗: {e}")
                
                self._price_cache = result
                
                # 合併 TPEx (上櫃) 資料
                tpex_data = self._fetch_tpex_prices()
                result.update(tpex_data)
                self._price_cache = result
                
                return result
                
            except Exception as e:
                last_error = e
                if attempt < MAX_RETRIES:
                    wait = attempt * 3
                    print(f"[WARNING] 抓取股價資料失敗 (第{attempt}次): {e}，{wait}秒後重試...")
                    time.sleep(wait)
                else:
                    print(f"[WARNING] 抓取股價資料失敗 (已重試{MAX_RETRIES}次): {last_error}")
        return {}

    def _fetch_tpex_prices(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 股價資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            # TPEx API: 資料在 tables[0]['data']
            tables = data.get('tables', [])
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 9:
                        code = str(row[0]).strip()
                        
                        def parse_float(val):
                            try:
                                clean = str(val).replace(',', '').replace('---', '0').replace('+', '').strip()
                                return float(clean) if clean else 0.0
                            except:
                                return 0.0
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', ''))
                            except:
                                return 0
                        
                        # TPEx 股價欄位:
                        # 0: 代號, 1: 名稱, 2: 收盤, 3: 漲跌, 4: 開盤, 
                        # 5: 最高, 6: 最低, 7: 均價, 8: 成交股數
                        close = parse_float(row[2])
                        change_val = parse_float(row[3])
                        volume = parse_int(row[8]) // 1000 if len(row) > 8 else 0
                        
                        pct_change = 0.0
                        if close > 0 and (close - change_val) != 0:
                            pct_change = round((change_val / (close - change_val)) * 100, 2)
                        
                        result[code] = {
                            'name': str(row[1]).strip(),
                            'market': '上櫃',
                            'close': close,
                            'change': change_val,
                            'pct_change': pct_change,
                            'volume': volume,
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 股價資料失敗: {e}")
            return {}
    
    def fetch_historical_prices(self, stock_code: str, days: int = 25) -> List[float]:
        """
        抓取個股歷史收盤價 (用於計算 MA)
        Returns:
            最近 N 天的收盤價列表
        """
        try:
            # 計算起始日期
            end_date = datetime.strptime(self.date_str, '%Y%m%d')
            start_date = end_date - timedelta(days=days + 15)  # 多抓一些以確保有足夠交易日
            
            url = f"{self.TWSE_BASE_URL}/exchangeReport/STOCK_DAY"
            params = {
                'response': 'json',
                'date': self.date_str,
                'stockNo': stock_code
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                return []
            
            prices = []
            for row in data.get('data', []):
                if len(row) >= 7:
                    try:
                        close = float(str(row[6]).replace(',', ''))
                        prices.append(close)
                    except:
                        pass
            
            return prices[-days:] if len(prices) >= days else prices
            
        except Exception as e:
            print(f"[WARNING] 抓取 {stock_code} 歷史價格失敗: {e}")
            return []


class StockMonitor:
    """個股籌碼監控主類別"""
    
    def __init__(self, date_str: str, watchlist_path: str = 'data/config/watchlist.json'):
        """
        Args:
            date_str: 日期字串，格式 YYYYMMDD
            watchlist_path: 自選股清單路徑
        """
        self.date_str = date_str
        self.watchlist_path = watchlist_path
        self.watchlist = []
        self.stock_data = {}
        self.warrant_data = {}
    
    def load_watchlist(self) -> List[Dict[str, str]]:
        """載入自選股清單"""
        try:
            with open(self.watchlist_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.watchlist = data.get('watchlist', [])
                print(f"[INFO] 已載入 {len(self.watchlist)} 檔自選股")
                return self.watchlist
        except FileNotFoundError:
            print(f"[ERROR] 找不到自選股清單: {self.watchlist_path}")
            return []
        except json.JSONDecodeError as e:
            print(f"[ERROR] 自選股清單格式錯誤: {e}")
            return []
    
    def fetch_all_data(self):
        """抓取所有個股資料"""
        if not self.watchlist:
            self.load_watchlist()
        
        if not self.watchlist:
            print("[ERROR] 沒有任何自選股")
            return
        
        print(f"\n[INFO] 開始抓取 {self.date_str} 的個股籌碼資料...\n")
        
        # 抓取當日資料
        print("[1/4] 抓取三大法人個股買賣超...")
        fetcher = StockDataFetcher(self.date_str)
        institutional = fetcher.fetch_institutional_trading()
        
        print("[2/4] 抓取融資融券餘額...")
        margin = fetcher.fetch_margin_trading()
        
        print("[3/4] 抓取個股收盤行情...")
        prices = fetcher.fetch_stock_prices()
        
        # 抓取過去 5 天資料計算累計
        print("[4/4] 計算 5 日累計數據...")
        history_data = self._fetch_5d_history()
        
        # 整合資料
        for stock in self.watchlist:
            code = stock['code']
            name = stock['name']
            
            inst_data = institutional.get(code, {})
            margin_data = margin.get(code, {})
            price_data = prices.get(code, {})
            hist = history_data.get(code, {})
            
            # 如果自選股沒填名稱，嘗試從獲取的資料補齊
            if not name:
                name = price_data.get('name', '') or inst_data.get('name', '')
            
            # 判斷是否為權證 (長度大於等於 6 碼)
            is_warrant = len(code) >= 6
            
            if is_warrant:
                # 取得權證基本資料 (履約價、行使比例、到期日)
                master = fetcher.fetch_warrant_master().get(code, {})
                
                # 計算剩餘天數
                days_to_expiry = None
                if master.get('expiry_date'):
                    try:
                        exp = datetime.strptime(master['expiry_date'], '%Y%m%d')
                        now = datetime.strptime(self.date_str, '%Y%m%d')
                        days_to_expiry = (exp - now).days
                    except:
                        pass
                
                # 計算價內外程度 (Underlying / Strike - 1)
                moneyness = None
                strike = master.get('strike_price', 0)
                underlying_p = price_data.get('underlying_price', 0)
                if strike > 0 and underlying_p > 0:
                    moneyness = f"{round((underlying_p / strike - 1) * 100, 2)}%"
                    if underlying_p > strike:
                        moneyness += " 價內"
                    else:
                        moneyness += " 價外"

                # 計算買賣價差比 (Ask-Bid)/((Ask+Bid)/2)
                spread_pct = None
                bid = price_data.get('bid', 0)
                ask = price_data.get('ask', 0)
                if ask > 0 and bid > 0:
                    spread_pct = round(((ask - bid) / ((ask + bid) / 2)) * 100, 2)

                # 計算名目槓桿 (標的價 / 權證價 * 行使比例)
                leverage = None
                warrant_p = price_data.get('close', 0)
                ratio = master.get('exercise_ratio', 0)
                if warrant_p > 0 and underlying_p > 0 and ratio > 0:
                    leverage = round((underlying_p / warrant_p) * ratio, 2)

                self.warrant_data[code] = {
                    'code': code,
                    'name': name,
                    'close': price_data.get('close'),
                    'change': price_data.get('change'),
                    'pct_change': price_data.get('pct_change'),
                    'volume': price_data.get('volume'),
                    'strike_price': strike if strike > 0 else None,
                    'exercise_ratio': ratio if ratio > 0 else None,
                    'days_to_expiry': days_to_expiry,
                    'moneyness': moneyness,
                    'bid_ask_spread_pct': spread_pct,
                    'effective_leverage': leverage, # 此處暫以名目槓桿代替
                    'implied_volatility': None, # 無法計算
                    'outstanding_pct': None # 無法從 API 取得
                }
                continue

            # 計算 MA20 乖離率
            dist_ma20 = None
            if price_data.get('close') and hist.get('ma20'):
                ma20 = hist['ma20']
                close = price_data['close']
                dist_ma20 = round((close - ma20) / ma20 * 100, 2)
            
            self.stock_data[code] = {
                'code': code,
                'name': name,
                'market': price_data.get('market', ''),
                
                # 價格資料
                'close': price_data.get('close'),
                'pct_change': price_data.get('pct_change'),
                'volume': price_data.get('volume'),
                
                # 三大法人當日 (轉為張)
                'foreign_daily': round(inst_data.get('foreign_net', 0) / 1000, 0) if inst_data.get('foreign_net') else None,
                'trust_daily': round(inst_data.get('trust_net', 0) / 1000, 0) if inst_data.get('trust_net') else None,
                'dealer_daily': round(inst_data.get('dealer_net', 0) / 1000, 0) if inst_data.get('dealer_net') else None,
                
                # 三大法人 5 日累計
                'foreign_5d_sum': hist.get('foreign_5d_sum'),
                'trust_5d_sum': hist.get('trust_5d_sum'),
                'dealer_5d_sum': hist.get('dealer_5d_sum'),
                
                # 融資融券
                'margin_daily_change': margin_data.get('margin_change'),
                'margin_5d_sum': hist.get('margin_5d_sum'),
                'lending_daily_change': margin_data.get('short_change'),
                
                # 技術面
                'dist_ma20': dist_ma20,
                
                # 5日成交量 (用於進階指標計算)
                'volume_5d': hist.get('volume_5d', 0),
            }
        
        print(f"[DEBUG] self.warrant_data size: {len(self.warrant_data)}")
        if '055145' in self.warrant_data:
            print(f"[DEBUG] 055145 Warrant Data: {self.warrant_data['055145']}")
        else:
            print("[DEBUG] 055145 NOT found in self.warrant_data")
            
        print("\n[SUCCESS] 個股籌碼資料抓取完成！\n")
    
    def _fetch_5d_history(self) -> Dict[str, Dict[str, Any]]:
        """抓取過去 5 天資料計算累計"""
        result = {}
        
        # Include the report date in the rolling 5-day window.
        candidate_days = get_previous_trading_days(self.date_str, 5, buffer_days=10)
        # 收集歷史資料
        history_institutional = {}
        history_margin = {}
        history_prices = {}
        
        valid_days_count = 0
        
        for date in candidate_days:
            if valid_days_count >= 5:
                break
                
            print(f"    嘗試抓取 {date} 歷史資料... (已收集 {valid_days_count}/5)")
            fetcher = StockDataFetcher(date)
            
            # 以股價資料是否存在來判斷是否為真實開市日
            prices = fetcher.fetch_stock_prices(include_warrants=False)
            if not prices:
                print(f"    - {date} 無資料 (可能是假日)，跳過")
                time.sleep(1.0)
                continue
                
            valid_days_count += 1
            
            for code, data in prices.items():
                if code not in history_prices:
                    history_prices[code] = []
                history_prices[code].append({
                    'close': data.get('close', 0),
                    'volume': data.get('volume', 0)
                })
            
            inst = fetcher.fetch_institutional_trading()
            for code, data in inst.items():
                if code not in history_institutional:
                    history_institutional[code] = []
                history_institutional[code].append(data)
            
            margin = fetcher.fetch_margin_trading()
            for code, data in margin.items():
                if code not in history_margin:
                    history_margin[code] = []
                history_margin[code].append(data)
            
            time.sleep(2.0)  # 避免請求太頻繁 (TWSE API 限制)
        
        # 計算累計值
        for stock in self.watchlist:
            code = stock['code']
            
            # 三大法人 5 日累計
            foreign_5d = 0
            trust_5d = 0
            dealer_5d = 0
            
            if code in history_institutional:
                for data in history_institutional[code]:
                    foreign_5d += data.get('foreign_net', 0) / 1000  # 轉張
                    trust_5d += data.get('trust_net', 0) / 1000
                    dealer_5d += data.get('dealer_net', 0) / 1000
            
            # 融資 5 日累計
            margin_5d = 0
            if code in history_margin:
                for data in history_margin[code]:
                    margin_5d += data.get('margin_change', 0)
            
            # 計算 MA20 和 5 日成交量
            ma20 = None
            volume_5d = 0
            if code in history_prices and len(history_prices[code]) >= 5:
                # 需要更多歷史資料來計算 MA20，這裡先簡化
                price_list = [p['close'] for p in history_prices[code] if p['close']]
                if len(price_list) >= 5:
                    ma20 = sum(price_list) / len(price_list)  # 簡化計算
                
                # 5 日成交量總和
                volume_5d = sum(p.get('volume', 0) for p in history_prices[code])
            
            result[code] = {
                'foreign_5d_sum': round(foreign_5d, 0),
                'trust_5d_sum': round(trust_5d, 0),
                'dealer_5d_sum': round(dealer_5d, 0),
                'margin_5d_sum': round(margin_5d, 0),
                'ma20': ma20,
                'volume_5d': volume_5d,
            }
        
        return result
    
    def display(self):
        """以表格形式顯示資料"""
        from tabulate import tabulate
        
        print(f"{'='*100}")
        print(f" 個股籌碼監控報告 - {self.date_str}")
        print(f"{'='*100}\n")
        
        table_data = []
        for code, data in self.stock_data.items():
            def fmt(val, suffix=''):
                if val is None:
                    return 'N/A'
                if isinstance(val, float):
                    return f"{val:+,.0f}{suffix}" if val != 0 else f"0{suffix}"
                return f"{val:+,}{suffix}" if isinstance(val, int) and val != 0 else str(val)
            
            table_data.append([
                data['code'],
                data['name'],
                data.get('market', ''),
                f"{data['close']:.2f}" if data['close'] else 'N/A',
                f"{data['pct_change']:+.2f}%" if data['pct_change'] else 'N/A',
                fmt(data['volume']),
                fmt(data['foreign_daily']),
                fmt(data['foreign_5d_sum']),
                fmt(data['trust_daily']),
                fmt(data['trust_5d_sum']),
                fmt(data['margin_daily_change']),
                fmt(data['margin_5d_sum']),
                f"{data['dist_ma20']:+.2f}%" if data['dist_ma20'] else 'N/A',
            ])
        
        headers = ['代號', '名稱', '市場別', '收盤價', '漲跌幅', '成交量', 
                   '外資(張)', '外資5日', '投信(張)', '投信5日',
                   '融資增減', '融資5日', 'MA20乖離']
        
        print(tabulate(table_data, headers=headers, tablefmt='grid'))
        print()
    
    def export_to_excel(self, filename: str = None):
        """匯出到 Excel 檔案"""
        if filename is None:
            filename = f"stock_monitor_{self.date_str}.xlsx"
        
        print(f"[INFO] 正在生成 Excel 報表...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "個股籌碼監控"
        
        # 標題
        ws['A1'] = f"個股籌碼監控報告 - {self.date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表頭
        headers = [
            '股票代號', '股票名稱', '市場別', '收盤價', '漲跌幅(%)', '成交量(張)',
            '外資當日(張)', '外資5日累計', '投信當日(張)', '投信5日累計', '自營商當日(張)',
            '融資增減(張)', '融資5日累計', '借券增減(張)', 'MA20乖離(%)'
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 資料行
        row = 4
        for code, data in self.stock_data.items():
            ws.cell(row, 1, data['code'])
            ws.cell(row, 2, data['name'])
            ws.cell(row, 3, data.get('market', ''))
            ws.cell(row, 4, data['close'])
            ws.cell(row, 5, data['pct_change'])
            ws.cell(row, 6, data['volume'])
            ws.cell(row, 7, data['foreign_daily'])
            ws.cell(row, 8, data['foreign_5d_sum'])
            ws.cell(row, 9, data['trust_daily'])
            ws.cell(row, 10, data['trust_5d_sum'])
            ws.cell(row, 11, data['dealer_daily'])
            ws.cell(row, 12, data['margin_daily_change'])
            ws.cell(row, 13, data['margin_5d_sum'])
            ws.cell(row, 14, data['lending_daily_change'])
            ws.cell(row, 15, data['dist_ma20'])
            
            # 條件格式 - 外資買超綠色，賣超紅色
            if data['foreign_daily'] and data['foreign_daily'] > 0:
                ws.cell(row, 7).font = Font(color="008000")
            elif data['foreign_daily'] and data['foreign_daily'] < 0:
                ws.cell(row, 7).font = Font(color="FF0000")
            
            row += 1
        
        # 調整欄寬 (含進階指標欄位)
        col_widths = [10, 12, 10, 10, 10, 12, 14, 14, 14, 14, 14, 12, 12, 12, 12]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 儲存
        output_dir = 'monitor_xlsx'
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        wb.save(output_path)
        print(f"[SUCCESS] Excel 報表已儲存至: {output_path}\n")
    
    def export_to_csv(self, filename: str = None):
        """匯出到 CSV 檔案 (含進階籌碼指標)"""
        if filename is None:
            filename = f"stock_monitor_{self.date_str}.csv"
        
        print(f"[INFO] 正在生成 CSV 報表...")
        
        # 準備資料
        rows = []
        for code, data in self.stock_data.items():
            rows.append({
                '日期': self.date_str,
                '股票代號': data.get('code', ''),
                '股票名稱': data.get('name', ''),
                '市場別': data.get('market', ''),
                '收盤價': data.get('close'),
                '成交量(張)': data.get('volume'),
                '外資當日(張)': data.get('foreign_daily'),
                '外資5日累計': data.get('foreign_5d_sum'),
                '投信當日(張)': data.get('trust_daily'),
                '投信5日累計': data.get('trust_5d_sum'),
                '自營商當日(張)': data.get('dealer_daily'),
                '融資增減(張)': data.get('margin_daily_change'),
                '融資5日累計': data.get('margin_5d_sum'),
                '借券增減(張)': data.get('lending_daily_change'),
                'MA20乖離(%)': data.get('dist_ma20'),
            })
        
        df = pd.DataFrame(rows)
        
        # 儲存
        output_dir = 'monitor_xlsx'
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"[SUCCESS] CSV 報表已儲存至: {output_path}\n")
    



def main():
    """主程式"""
    parser = argparse.ArgumentParser(
        description='台灣個股籌碼監控程式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例:
  # 抓取今天的個股籌碼資料
  python stock_monitor.py
  
  # 指定日期
  python stock_monitor.py --date 20260203
  
  # 指定自選股清單
  python stock_monitor.py --watchlist my_stocks.json
        """
    )
    
    parser.add_argument(
        '--date',
        type=str,
        help='指定日期 (格式: YYYYMMDD)，不指定則使用今天'
    )
    
    parser.add_argument(
        '--watchlist',
        type=str,
        default='data/config/watchlist.json',
        help='自選股清單路徑 (預設: data/config/watchlist.json)'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        help='Excel 輸出檔名'
    )
    

    
    parser.add_argument(
        '--csv',
        action='store_true',
        help='同時輸出 CSV 檔案'
    )
    
    args = parser.parse_args()
    
    # 取得交易日期
    trading_date = get_trading_date(args.date)
    
    if args.date and args.date != trading_date:
        print(f"[INFO] 指定日期為週末，已調整為 {trading_date}\n")
    
    # 執行監控
    try:
        monitor = StockMonitor(trading_date, args.watchlist)
        monitor.load_watchlist()
        monitor.fetch_all_data()
        monitor.display()
        
        output_filename = args.output or f"stock_monitor_{trading_date}.xlsx"
        monitor.export_to_excel(output_filename)
        
        # 輸出 CSV
        if args.csv:
            csv_filename = output_filename.replace('.xlsx', '.csv')
            monitor.export_to_csv(csv_filename)
        
        print(f"[SUCCESS] 完成！報告已儲存至 monitor_xlsx/{output_filename}")
        
    except KeyboardInterrupt:
        print("\n\n[WARNING] 程式被使用者中斷")
        return 1
    except Exception as e:
        print(f"\n[ERROR] 執行失敗: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())
