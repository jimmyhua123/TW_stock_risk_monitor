#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
台灣個股籌碼監控程式
抓取個股三大法人買賣超、融資融券、技術面資料
"""

import requests
import pandas as pd
import json
import os
import argparse
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any, List, Optional
from io import StringIO
import time

# 匯入現有模組
# 匯入現有模組
from risk_monitor import get_trading_date


class StockDataFetcher:
    """個股資料抓取器 (支援上市 TWSE + 上櫃 TPEx)"""
    
    TWSE_BASE_URL = "https://www.twse.com.tw"
    TPEX_BASE_URL = "https://www.tpex.org.tw"
    TAIFEX_BASE_URL = "https://www.taifex.com.tw"
    
    def __init__(self, date_str: str):
        """
        Args:
            date_str: 日期字串，格式 YYYYMMDD
        """
        self.date_str = date_str
        self.formatted_date = f"{date_str[:4]}/{date_str[4:6]}/{date_str[6:]}"
        # TPEx 使用民國年格式: YYY/MM/DD
        year_ad = int(date_str[:4])
        year_roc = year_ad - 1911
        self.tpex_date = f"{year_roc}/{date_str[4:6]}/{date_str[6:]}"
        
        # 快取資料
        self._institutional_cache = None
        self._margin_cache = None
        self._price_cache = None
    
    def fetch_institutional_trading(self) -> Dict[str, Dict[str, Any]]:
        """
        抓取三大法人個股買賣超資料 (T86)
        Returns:
            {
                '2330': {
                    'foreign_buy': 外資買進股數,
                    'foreign_sell': 外資賣出股數,
                    'foreign_net': 外資買賣超股數,
                    'trust_buy': 投信買進股數,
                    'trust_sell': 投信賣出股數,
                    'trust_net': 投信買賣超股數,
                    'dealer_buy': 自營商買進股數,
                    'dealer_sell': 自營商賣出股數,
                    'dealer_net': 自營商買賣超股數,
                },
                ...
            }
        """
        if self._institutional_cache is not None:
            return self._institutional_cache
            
        try:
            url = f"{self.TWSE_BASE_URL}/fund/T86"
            params = {
                'response': 'json',
                'date': self.date_str,
                'selectType': 'ALLBUT0999'  # 全部但排除權證
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                print(f"[WARNING] T86 API 返回錯誤: {data.get('stat')}")
                return {}
            
            result = {}
            for row in data.get('data', []):
                if len(row) >= 17:
                    code = row[0].strip()
                    name = row[1].strip()
                    
                    # 解析數值 (移除逗號)
                    def parse_int(val):
                        try:
                            return int(str(val).replace(',', ''))
                        except:
                            return 0
                    
                    # T86 欄位順序:
                    # 0: 證券代號, 1: 證券名稱
                    # 2-4: 外陸資買/賣/淨 (不含自營商)
                    # 5-7: 外資自營商買/賣/淨
                    # 8-10: 投信買/賣/淨
                    # 11-13: 自營商(自行)買/賣/淨
                    # 14-16: 自營商(避險)買/賣/淨
                    # 17: 三大法人合計買賣超
                    result[code] = {
                        'name': name,
                        'foreign_buy': parse_int(row[2]),
                        'foreign_sell': parse_int(row[3]),
                        'foreign_net': parse_int(row[4]),
                        'trust_buy': parse_int(row[8]),
                        'trust_sell': parse_int(row[9]),
                        'trust_net': parse_int(row[10]),
                        'dealer_buy': parse_int(row[11]) + parse_int(row[14]),  # 自行買賣 + 避險
                        'dealer_sell': parse_int(row[12]) + parse_int(row[15]),
                        'dealer_net': parse_int(row[13]) + parse_int(row[16]),
                    }
            
            self._institutional_cache = result
            
            # 合併 TPEx (上櫃) 資料
            tpex_data = self._fetch_tpex_institutional()
            result.update(tpex_data)
            self._institutional_cache = result
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取三大法人個股資料失敗: {e}")
            return {}
    
    def _fetch_tpex_institutional(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 三大法人買賣超資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/3insti/daily_trade/3itrade_hedge_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            tables = data.get('tables', [])
            # TPEx API: tables[0] 是 dict，資料在 tables[0]['data']
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 24:
                        code = str(row[0]).strip()
                        name = str(row[1]).strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', ''))
                            except:
                                return 0
                        
                        # TPEx 欄位順序:
                        # 0: 代號, 1: 名稱
                        # 2-4: 外資買/賣/淨 (不含自營商)
                        # 5-7: 外資自營商買/賣/淨  
                        # 8-10: 外資合計買/賣/淨
                        # 11-13: 投信買/賣/淨
                        # 14-16: 自營商(自行)買/賣/淨
                        # 17-19: 自營商(避險)買/賣/淨
                        # 20-22: 自營商合計買/賣/淨
                        # 23: 三大法人合計
                        result[code] = {
                            'name': name,
                            'foreign_buy': parse_int(row[2]),
                            'foreign_sell': parse_int(row[3]),
                            'foreign_net': parse_int(row[4]),
                            'trust_buy': parse_int(row[11]),
                            'trust_sell': parse_int(row[12]),
                            'trust_net': parse_int(row[13]),
                            'dealer_buy': parse_int(row[14]) + parse_int(row[17]),
                            'dealer_sell': parse_int(row[15]) + parse_int(row[18]),
                            'dealer_net': parse_int(row[16]) + parse_int(row[19]),
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 三大法人資料失敗: {e}")
            return {}
    
    def fetch_margin_trading(self) -> Dict[str, Dict[str, Any]]:
        """
        抓取個股融資融券餘額
        Returns:
            {
                '2330': {
                    'margin_buy': 融資買進,
                    'margin_sell': 融資賣出,
                    'margin_balance': 融資餘額,
                    'margin_change': 融資增減,
                    'short_buy': 融券買進,
                    'short_sell': 融券賣出,
                    'short_balance': 融券餘額,
                    'short_change': 融券增減,
                },
                ...
            }
        """
        if self._margin_cache is not None:
            return self._margin_cache
            
        try:
            url = f"{self.TWSE_BASE_URL}/rwd/zh/marginTrading/MI_MARGN"
            params = {
                'response': 'json',
                'date': self.date_str,
                'selectType': 'ALL'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                print(f"[WARNING] 融資融券 API 返回錯誤: {data.get('stat')}")
                return {}
            
            result = {}
            
            # 解析個股融資融券資料 (tables[1] 是個股明細)
            if len(data.get('tables', [])) > 1:
                table_data = data['tables'][1].get('data', [])
                
                for row in table_data:
                    if len(row) >= 13:
                        code = row[0].strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', '').replace('--', '0'))
                            except:
                                return 0
                        
                        result[code] = {
                            'margin_buy': parse_int(row[2]),
                            'margin_sell': parse_int(row[3]),
                            'margin_balance': parse_int(row[6]),
                            'margin_change': parse_int(row[6]) - parse_int(row[1]) if row[1] != '--' else 0,
                            'short_sell': parse_int(row[8]),
                            'short_buy': parse_int(row[9]),
                            'short_balance': parse_int(row[12]),
                            'short_change': parse_int(row[12]) - parse_int(row[7]) if row[7] != '--' else 0,
                        }
            
            self._margin_cache = result
            
            # 合併 TPEx (上櫃) 資料
            tpex_data = self._fetch_tpex_margin()
            result.update(tpex_data)
            self._margin_cache = result
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取融資融券資料失敗: {e}")
            return {}
    
    def _fetch_tpex_margin(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 融資融券資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/margin_trading/margin_balance/margin_bal_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            # TPEx API: 資料在 tables[0]['data']
            tables = data.get('tables', [])
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 7:
                        code = str(row[0]).strip()
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', '').replace('--', '0'))
                            except:
                                return 0
                        
                        # TPEx 融資融券欄位:
                        # 0: 代號, 1: 名稱
                        # 2: 前日融資餘額, 3: 融資買進, 4: 融資賣出, 5: 現金償還, 6: 今日融資餘額
                        result[code] = {
                            'margin_buy': parse_int(row[3]) if len(row) > 3 else 0,
                            'margin_sell': parse_int(row[4]) if len(row) > 4 else 0,
                            'margin_balance': parse_int(row[6]) if len(row) > 6 else 0,
                            'margin_change': parse_int(row[6]) - parse_int(row[2]) if len(row) > 6 else 0,
                            'short_sell': 0,
                            'short_buy': 0,
                            'short_balance': 0,
                            'short_change': 0,
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 融資融券資料失敗: {e}")
            return {}
    
    def fetch_stock_prices(self) -> Dict[str, Dict[str, Any]]:
        """
        抓取個股收盤行情
        Returns:
            {
                '2330': {
                    'close': 收盤價,
                    'change': 漲跌,
                    'pct_change': 漲跌幅 %,
                    'volume': 成交量 (張),
                },
                ...
            }
        """
        if self._price_cache is not None:
            return self._price_cache
            
        try:
            # 使用新版 API 端點
            url = f"{self.TWSE_BASE_URL}/rwd/zh/afterTrading/MI_INDEX"
            params = {
                'response': 'json',
                'date': self.date_str,
                'type': 'ALLBUT0999'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                print(f"[WARNING] 股價 API 返回錯誤: {data.get('stat')}")
                return {}
            
            result = {}
            
            # 個股資料在 tables[8] - 每日收盤行情
            tables = data.get('tables', [])
            if len(tables) > 8:
                stock_table = tables[8]
                for row in stock_table.get('data', []):
                    if len(row) >= 11:
                        code = str(row[0]).strip()
                        
                        # 只處理數字代號
                        if not code.isdigit():
                            continue
                        
                        def parse_float(val):
                            try:
                                # 移除逗號、HTML標籤、其他符號
                                clean = str(val).replace(',', '').replace('X', '')
                                clean = clean.replace('<p style= color:green>', '').replace('<p style= color:red>', '').replace('</p>', '')
                                clean = clean.replace('+', '').strip()
                                if clean == '--' or clean == '':
                                    return 0.0
                                return float(clean)
                            except:
                                return 0.0
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', ''))
                            except:
                                return 0
                        
                        # 欄位: 0代號, 1名稱, 2成交股數, 3成交筆數, 4成交金額,
                        # 5開盤, 6最高, 7最低, 8收盤, 9漲跌方向, 10漲跌價差, 11-15其他
                        close = parse_float(row[8])
                        change_val = parse_float(row[10])
                        volume = parse_int(row[2]) // 1000  # 股數 -> 張
                        
                        # 判斷漲跌方向 (row[9] 包含 color:green 為跌)
                        if 'green' in str(row[9]):
                            change_val = -abs(change_val)
                        
                        pct_change = 0.0
                        if close > 0 and (close - change_val) != 0:
                            pct_change = round((change_val / (close - change_val)) * 100, 2)
                        
                        result[code] = {
                            'name': str(row[1]).strip(),
                            'close': close,
                            'change': change_val,
                            'pct_change': pct_change,
                            'volume': volume,
                        }
            
            self._price_cache = result
            
            # 合併 TPEx (上櫃) 資料
            tpex_data = self._fetch_tpex_prices()
            result.update(tpex_data)
            self._price_cache = result
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取股價資料失敗: {e}")
            return {}
    
    def _fetch_tpex_prices(self) -> Dict[str, Dict[str, Any]]:
        """抓取 TPEx (上櫃) 股價資料"""
        try:
            url = f"{self.TPEX_BASE_URL}/web/stock/aftertrading/daily_close_quotes/stk_quote_result.php"
            params = {
                'l': 'zh-tw',
                'o': 'json',
                'd': self.tpex_date,
                '_': '1'
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            result = {}
            # TPEx API: 資料在 tables[0]['data']
            tables = data.get('tables', [])
            if tables and len(tables) > 0 and isinstance(tables[0], dict):
                rows = tables[0].get('data', [])
                for row in rows:
                    if len(row) >= 9:
                        code = str(row[0]).strip()
                        
                        def parse_float(val):
                            try:
                                clean = str(val).replace(',', '').replace('---', '0').replace('+', '').strip()
                                return float(clean) if clean else 0.0
                            except:
                                return 0.0
                        
                        def parse_int(val):
                            try:
                                return int(str(val).replace(',', ''))
                            except:
                                return 0
                        
                        # TPEx 股價欄位:
                        # 0: 代號, 1: 名稱, 2: 收盤, 3: 漲跌, 4: 開盤, 
                        # 5: 最高, 6: 最低, 7: 均價, 8: 成交股數
                        close = parse_float(row[2])
                        change_val = parse_float(row[3])
                        volume = parse_int(row[8]) // 1000 if len(row) > 8 else 0
                        
                        pct_change = 0.0
                        if close > 0 and (close - change_val) != 0:
                            pct_change = round((change_val / (close - change_val)) * 100, 2)
                        
                        result[code] = {
                            'name': str(row[1]).strip(),
                            'close': close,
                            'change': change_val,
                            'pct_change': pct_change,
                            'volume': volume,
                        }
            
            return result
            
        except Exception as e:
            print(f"[WARNING] 抓取 TPEx 股價資料失敗: {e}")
            return {}
    
    def fetch_historical_prices(self, stock_code: str, days: int = 25) -> List[float]:
        """
        抓取個股歷史收盤價 (用於計算 MA)
        Returns:
            最近 N 天的收盤價列表
        """
        try:
            # 計算起始日期
            end_date = datetime.strptime(self.date_str, '%Y%m%d')
            start_date = end_date - timedelta(days=days + 15)  # 多抓一些以確保有足夠交易日
            
            url = f"{self.TWSE_BASE_URL}/exchangeReport/STOCK_DAY"
            params = {
                'response': 'json',
                'date': self.date_str,
                'stockNo': stock_code
            }
            
            response = requests.get(url, params=params, timeout=15)
            response.raise_for_status()
            data = response.json()
            
            if data.get('stat') != 'OK':
                return []
            
            prices = []
            for row in data.get('data', []):
                if len(row) >= 7:
                    try:
                        close = float(str(row[6]).replace(',', ''))
                        prices.append(close)
                    except:
                        pass
            
            return prices[-days:] if len(prices) >= days else prices
            
        except Exception as e:
            print(f"[WARNING] 抓取 {stock_code} 歷史價格失敗: {e}")
            return []


class StockMonitor:
    """個股籌碼監控主類別"""
    
    def __init__(self, date_str: str, watchlist_path: str = 'watchlist.json'):
        """
        Args:
            date_str: 日期字串，格式 YYYYMMDD
            watchlist_path: 自選股清單路徑
        """
        self.date_str = date_str
        self.watchlist_path = watchlist_path
        self.watchlist = []
        self.stock_data = {}
    
    def load_watchlist(self) -> List[Dict[str, str]]:
        """載入自選股清單"""
        try:
            with open(self.watchlist_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                self.watchlist = data.get('watchlist', [])
                print(f"[INFO] 已載入 {len(self.watchlist)} 檔自選股")
                return self.watchlist
        except FileNotFoundError:
            print(f"[ERROR] 找不到自選股清單: {self.watchlist_path}")
            return []
        except json.JSONDecodeError as e:
            print(f"[ERROR] 自選股清單格式錯誤: {e}")
            return []
    
    def fetch_all_data(self):
        """抓取所有個股資料"""
        if not self.watchlist:
            self.load_watchlist()
        
        if not self.watchlist:
            print("[ERROR] 沒有任何自選股")
            return
        
        print(f"\n[INFO] 開始抓取 {self.date_str} 的個股籌碼資料...\n")
        
        # 抓取當日資料
        print("[1/4] 抓取三大法人個股買賣超...")
        fetcher = StockDataFetcher(self.date_str)
        institutional = fetcher.fetch_institutional_trading()
        
        print("[2/4] 抓取融資融券餘額...")
        margin = fetcher.fetch_margin_trading()
        
        print("[3/4] 抓取個股收盤行情...")
        prices = fetcher.fetch_stock_prices()
        
        # 抓取過去 5 天資料計算累計
        print("[4/4] 計算 5 日累計數據...")
        history_data = self._fetch_5d_history()
        
        # 整合資料
        for stock in self.watchlist:
            code = stock['code']
            name = stock['name']
            
            inst_data = institutional.get(code, {})
            margin_data = margin.get(code, {})
            price_data = prices.get(code, {})
            hist = history_data.get(code, {})
            
            # 計算 MA20 乖離率
            dist_ma20 = None
            if price_data.get('close') and hist.get('ma20'):
                ma20 = hist['ma20']
                close = price_data['close']
                dist_ma20 = round((close - ma20) / ma20 * 100, 2)
            
            self.stock_data[code] = {
                'code': code,
                'name': name,
                
                # 價格資料
                'close': price_data.get('close'),
                'pct_change': price_data.get('pct_change'),
                'volume': price_data.get('volume'),
                
                # 三大法人當日 (轉為張)
                'foreign_daily': round(inst_data.get('foreign_net', 0) / 1000, 0) if inst_data.get('foreign_net') else None,
                'trust_daily': round(inst_data.get('trust_net', 0) / 1000, 0) if inst_data.get('trust_net') else None,
                'dealer_daily': round(inst_data.get('dealer_net', 0) / 1000, 0) if inst_data.get('dealer_net') else None,
                
                # 三大法人 5 日累計
                'foreign_5d_sum': hist.get('foreign_5d_sum'),
                'trust_5d_sum': hist.get('trust_5d_sum'),
                'dealer_5d_sum': hist.get('dealer_5d_sum'),
                
                # 融資融券
                'margin_daily_change': margin_data.get('margin_change'),
                'margin_5d_sum': hist.get('margin_5d_sum'),
                'lending_daily_change': margin_data.get('short_change'),
                
                # 技術面
                'dist_ma20': dist_ma20,
                
                # 5日成交量 (用於進階指標計算)
                'volume_5d': hist.get('volume_5d', 0),
            }
        
        print("\n[SUCCESS] 個股籌碼資料抓取完成！\n")
    
    def _fetch_5d_history(self) -> Dict[str, Dict[str, Any]]:
        """抓取過去 5 天資料計算累計"""
        result = {}
        
        # 計算過去 5 個交易日
        base_date = datetime.strptime(self.date_str, '%Y%m%d')
        trading_days = []
        check_date = base_date - timedelta(days=1)
        
        while len(trading_days) < 5:
            # 跳過週末
            if check_date.weekday() < 5:
                trading_days.append(check_date.strftime('%Y%m%d'))
            check_date -= timedelta(days=1)
        
        # 收集歷史資料
        history_institutional = {}
        history_margin = {}
        history_prices = {}
        
        for i, date in enumerate(trading_days):
            print(f"    抓取 {date} 歷史資料... ({i+1}/5)")
            fetcher = StockDataFetcher(date)
            
            inst = fetcher.fetch_institutional_trading()
            for code, data in inst.items():
                if code not in history_institutional:
                    history_institutional[code] = []
                history_institutional[code].append(data)
            
            margin = fetcher.fetch_margin_trading()
            for code, data in margin.items():
                if code not in history_margin:
                    history_margin[code] = []
                history_margin[code].append(data)
            
            prices = fetcher.fetch_stock_prices()
            for code, data in prices.items():
                if code not in history_prices:
                    history_prices[code] = []
                history_prices[code].append({
                    'close': data.get('close', 0),
                    'volume': data.get('volume', 0)
                })
            
            time.sleep(2.0)  # 避免請求太頻繁 (TWSE API 限制)
        
        # 計算累計值
        for stock in self.watchlist:
            code = stock['code']
            
            # 三大法人 5 日累計
            foreign_5d = 0
            trust_5d = 0
            dealer_5d = 0
            
            if code in history_institutional:
                for data in history_institutional[code]:
                    foreign_5d += data.get('foreign_net', 0) / 1000  # 轉張
                    trust_5d += data.get('trust_net', 0) / 1000
                    dealer_5d += data.get('dealer_net', 0) / 1000
            
            # 融資 5 日累計
            margin_5d = 0
            if code in history_margin:
                for data in history_margin[code]:
                    margin_5d += data.get('margin_change', 0)
            
            # 計算 MA20 和 5 日成交量
            ma20 = None
            volume_5d = 0
            if code in history_prices and len(history_prices[code]) >= 5:
                # 需要更多歷史資料來計算 MA20，這裡先簡化
                price_list = [p['close'] for p in history_prices[code] if p['close']]
                if len(price_list) >= 5:
                    ma20 = sum(price_list) / len(price_list)  # 簡化計算
                
                # 5 日成交量總和
                volume_5d = sum(p.get('volume', 0) for p in history_prices[code])
            
            result[code] = {
                'foreign_5d_sum': round(foreign_5d, 0),
                'trust_5d_sum': round(trust_5d, 0),
                'dealer_5d_sum': round(dealer_5d, 0),
                'margin_5d_sum': round(margin_5d, 0),
                'ma20': ma20,
                'volume_5d': volume_5d,
            }
        
        return result
    
    def display(self):
        """以表格形式顯示資料"""
        from tabulate import tabulate
        
        print(f"{'='*100}")
        print(f" 個股籌碼監控報告 - {self.date_str}")
        print(f"{'='*100}\n")
        
        table_data = []
        for code, data in self.stock_data.items():
            def fmt(val, suffix=''):
                if val is None:
                    return 'N/A'
                if isinstance(val, float):
                    return f"{val:+,.0f}{suffix}" if val != 0 else f"0{suffix}"
                return f"{val:+,}{suffix}" if isinstance(val, int) and val != 0 else str(val)
            
            table_data.append([
                data['code'],
                data['name'],
                f"{data['close']:.2f}" if data['close'] else 'N/A',
                f"{data['pct_change']:+.2f}%" if data['pct_change'] else 'N/A',
                fmt(data['volume']),
                fmt(data['foreign_daily']),
                fmt(data['foreign_5d_sum']),
                fmt(data['trust_daily']),
                fmt(data['trust_5d_sum']),
                fmt(data['margin_daily_change']),
                fmt(data['margin_5d_sum']),
                f"{data['dist_ma20']:+.2f}%" if data['dist_ma20'] else 'N/A',
            ])
        
        headers = ['代號', '名稱', '收盤價', '漲跌幅', '成交量', 
                   '外資(張)', '外資5日', '投信(張)', '投信5日',
                   '融資增減', '融資5日', 'MA20乖離']
        
        print(tabulate(table_data, headers=headers, tablefmt='grid'))
        print()
    
    def export_to_excel(self, filename: str = None):
        """匯出到 Excel 檔案"""
        if filename is None:
            filename = f"stock_monitor_{self.date_str}.xlsx"
        
        print(f"[INFO] 正在生成 Excel 報表...")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "個股籌碼監控"
        
        # 標題
        ws['A1'] = f"個股籌碼監控報告 - {self.date_str}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:O1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 表頭
        headers = [
            '股票代號', '股票名稱', '收盤價', '漲跌幅(%)', '成交量(張)',
            '外資當日(張)', '外資5日累計', '投信當日(張)', '投信5日累計', '自營商當日(張)',
            '融資增減(張)', '融資5日累計', '借券增減(張)', 'MA20乖離(%)', '籌碼評價'
        ]
        
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 資料行
        row = 4
        for code, data in self.stock_data.items():
            # 計算籌碼評價
            chips_score = self._evaluate_chips(data)
            
            ws.cell(row, 1, data['code'])
            ws.cell(row, 2, data['name'])
            ws.cell(row, 3, data['close'])
            ws.cell(row, 4, data['pct_change'])
            ws.cell(row, 5, data['volume'])
            ws.cell(row, 6, data['foreign_daily'])
            ws.cell(row, 7, data['foreign_5d_sum'])
            ws.cell(row, 8, data['trust_daily'])
            ws.cell(row, 9, data['trust_5d_sum'])
            ws.cell(row, 10, data['dealer_daily'])
            ws.cell(row, 11, data['margin_daily_change'])
            ws.cell(row, 12, data['margin_5d_sum'])
            ws.cell(row, 13, data['lending_daily_change'])
            ws.cell(row, 14, data['dist_ma20'])
            ws.cell(row, 15, chips_score)
            
            # 條件格式 - 外資買超綠色，賣超紅色
            if data['foreign_daily'] and data['foreign_daily'] > 0:
                ws.cell(row, 6).font = Font(color="008000")
            elif data['foreign_daily'] and data['foreign_daily'] < 0:
                ws.cell(row, 6).font = Font(color="FF0000")
            
            row += 1
        
        # 調整欄寬 (含進階指標欄位)
        col_widths = [10, 12, 10, 10, 12, 14, 14, 14, 14, 14, 12, 12, 12, 12, 12]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # 儲存
        output_dir = 'monitor_xlsx'
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        wb.save(output_path)
        print(f"[SUCCESS] Excel 報表已儲存至: {output_path}\n")
    
    def export_to_csv(self, filename: str = None):
        """匯出到 CSV 檔案 (含進階籌碼指標)"""
        if filename is None:
            filename = f"stock_monitor_{self.date_str}.csv"
        
        print(f"[INFO] 正在生成 CSV 報表...")
        
        # 準備資料
        rows = []
        for code, data in self.stock_data.items():
            chips_score = self._evaluate_chips(data)
            
            rows.append({
                '日期': self.date_str,
                '股票代號': data.get('code', ''),
                '股票名稱': data.get('name', ''),
                '收盤價': data.get('close'),
                '成交量(張)': data.get('volume'),
                '外資當日(張)': data.get('foreign_daily'),
                '外資5日累計': data.get('foreign_5d_sum'),
                '投信當日(張)': data.get('trust_daily'),
                '投信5日累計': data.get('trust_5d_sum'),
                '自營商當日(張)': data.get('dealer_daily'),
                '融資增減(張)': data.get('margin_daily_change'),
                '融資5日累計': data.get('margin_5d_sum'),
                '借券增減(張)': data.get('lending_daily_change'),
                'MA20乖離(%)': data.get('dist_ma20'),
                '籌碼評價': chips_score,
            })
        
        df = pd.DataFrame(rows)
        
        # 儲存
        output_dir = 'monitor_xlsx'
        os.makedirs(output_dir, exist_ok=True)
        output_path = os.path.join(output_dir, filename)
        
        df.to_csv(output_path, index=False, encoding='utf-8-sig')
        print(f"[SUCCESS] CSV 報表已儲存至: {output_path}\n")
    
    def _evaluate_chips(self, data: Dict) -> str:
        """評估籌碼面"""
        score = 0
        
        # 外資連續買
        if data.get('foreign_5d_sum') and data['foreign_5d_sum'] > 0:
            score += 2
        elif data.get('foreign_5d_sum') and data['foreign_5d_sum'] < 0:
            score -= 2
        
        # 投信連續買
        if data.get('trust_5d_sum') and data['trust_5d_sum'] > 0:
            score += 1
        elif data.get('trust_5d_sum') and data['trust_5d_sum'] < 0:
            score -= 1
        
        # 融資減少 (籌碼乾淨)
        if data.get('margin_5d_sum') and data['margin_5d_sum'] < 0:
            score += 1
        elif data.get('margin_5d_sum') and data['margin_5d_sum'] > 0:
            score -= 1
        
        # 評價
        if score >= 3:
            return "主力積極買進"
        elif score >= 1:
            return "偏多"
        elif score <= -3:
            return "主力積極賣出"
        elif score <= -1:
            return "偏空"
        else:
            return "中性"


def main():
    """主程式"""
    parser = argparse.ArgumentParser(
        description='台灣個股籌碼監控程式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
範例:
  # 抓取今天的個股籌碼資料
  python stock_monitor.py
  
  # 指定日期
  python stock_monitor.py --date 20260203
  
  # 指定自選股清單
  python stock_monitor.py --watchlist my_stocks.json
        """
    )
    
    parser.add_argument(
        '--date',
        type=str,
        help='指定日期 (格式: YYYYMMDD)，不指定則使用今天'
    )
    
    parser.add_argument(
        '--watchlist',
        type=str,
        default='watchlist.json',
        help='自選股清單路徑 (預設: watchlist.json)'
    )
    
    parser.add_argument(
        '--output',
        type=str,
        help='Excel 輸出檔名'
    )
    

    
    parser.add_argument(
        '--csv',
        action='store_true',
        help='同時輸出 CSV 檔案'
    )
    
    args = parser.parse_args()
    
    # 取得交易日期
    trading_date = get_trading_date(args.date)
    
    if args.date and args.date != trading_date:
        print(f"[INFO] 指定日期為週末，已調整為 {trading_date}\n")
    
    # 執行監控
    try:
        monitor = StockMonitor(trading_date, args.watchlist)
        monitor.load_watchlist()
        monitor.fetch_all_data()
        monitor.display()
        
        output_filename = args.output or f"stock_monitor_{trading_date}.xlsx"
        monitor.export_to_excel(output_filename)
        
        # 輸出 CSV
        if args.csv:
            csv_filename = output_filename.replace('.xlsx', '.csv')
            monitor.export_to_csv(csv_filename)
        
        print(f"[SUCCESS] 完成！報告已儲存至 monitor_xlsx/{output_filename}")
        
    except KeyboardInterrupt:
        print("\n\n[WARNING] 程式被使用者中斷")
        return 1
    except Exception as e:
        print(f"\n[ERROR] 執行失敗: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())
